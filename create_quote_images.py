from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
import os
import textwrap


# Đường dẫn tới file Excel chứa danh sách quotes của bạn
quotes_file = 'C:/Users/Lan Huong/Desktop/Xoa duoc/0. GPT/Quotes-list.xlsx'

# Đường dẫn tới template của bạn
template_file = 'C:/Users/Lan Huong/Desktop/Xoa duoc/0. GPT/Template/11e-template.jpg'

# Đọc danh sách quotes từ file Excel
workbook = load_workbook(quotes_file)
worksheet = workbook.active

# Đường dẫn tới thư mục lưu trữ ảnh đã tạo
output_folder = 'C:/Users/Lan Huong/Desktop/Xoa duoc/0. GPT/Quotes Photo/'

# Khởi tạo kích thước của ảnh
image_size = (1920, 1080)

# Khởi tạo font chữ
font_path = os.path.join(os.getcwd(), 'fonts', 'Baskerville Bold font.ttf')
font_size = 52
font = ImageFont.truetype(font_path, font_size)

# Define line spacing and wrap property
line_spacing = 20
max_quote_width = 1785
quote_left = 65
wrap = True

# Lặp qua từng quote trong danh sách
for row in worksheet.iter_rows(min_row=2, values_only=True):
    quote = row[0]

    # Tạo ảnh mới bằng cách tải template lên
    image = Image.open(template_file).resize(image_size)
    draw = ImageDraw.Draw(image)

    # Tạo vùng chứa quote
    quote_max_width = max_quote_width
    quote_lines = textwrap.wrap(quote, width=quote_max_width//font_size, break_long_words=wrap)
    total_quote_height = len(quote_lines) * (font_size + line_spacing) - line_spacing
    quote_top = (image_size[1] - total_quote_height) // 2

    # Tạo các chuỗi con dựa trên độ rộng giới hạn của vùng chứa và line_spacing
    quote_lines = textwrap.wrap(quote, width=quote_max_width//font_size, break_long_words=wrap)

    # Tính toán vị trí y của từng dòng quote
    y = quote_top
    for line in quote_lines:
        line_width, line_height = draw.textsize(line, font=font)
        x = quote_left + (quote_max_width - line_width) // 2
        draw.text((x, y), line, font=font, fill="#9a8479", align='center')
        y += line_height + line_spacing

    # Lưu ảnh vào thư mục đầu ra
    output_file = output_folder + f'Quote-{str(row[0]).zfill(2)}.jpg'
    image.save(output_file)
